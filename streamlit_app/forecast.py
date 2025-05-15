
import streamlit as st
import torch
import torch.nn as nn
import pandas as pd
import numpy as np
import json
from datetime import datetime, timedelta
from io import StringIO, BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 参数配置
DATA_COLUMNS = ['evaporation_from_bare_soil_sum',
                'total_precipitation_sum',
                'temperature_2m_max',
                'wind_speed_10m']

HISTORY_DAYS = 15
FORECAST_DAYS = 7
INPUT_SIZE = len(DATA_COLUMNS)

# 加载模型参数
with open("models/best_params.json", "r") as f:
    best_params = json.load(f)

# 定义模型
class LSTMRunoffModel(nn.Module):
    def __init__(self, input_size, hidden_size1, hidden_size2, dropout=0.1):
        super().__init__()
        self.lstm1 = nn.LSTM(input_size=input_size, hidden_size=hidden_size1, batch_first=True)
        self.lstm2 = nn.LSTM(input_size=hidden_size1, hidden_size=hidden_size2, batch_first=True)
        self.fc = nn.Linear(hidden_size2, 1)
        self.dropout = nn.Dropout(dropout)

    def forward(self, x):
        out, _ = self.lstm1(x)
        out = self.dropout(out)
        out, _ = self.lstm2(out)
        out = self.fc(out)
        return out.squeeze(-1)

# 加载模型
@st.cache_resource
def load_model():
    model = LSTMRunoffModel(INPUT_SIZE, best_params['hidden_size1'], best_params['hidden_size2'], best_params['dropout'])
    model.load_state_dict(torch.load("models/best_lstm_model.pth", map_location="cpu"))
    model.eval()
    return model

# 标准化
def normalize_input(data):
    return (data - data.mean()) / (data.std() + 1e-8)

# 生成Excel模板
def create_excel_template():
    # 创建工作簿和工作表
    wb = openpyxl.Workbook()
    ws_data = wb.active
    ws_data.title = "数据输入"
    
    # 添加第二个工作表用于说明
    ws_guide = wb.create_sheet(title="填写指南")
    
    # ===== 工作表1: 数据输入区域 =====
    # 设置表头
    headers = ['date', 'evaporation_from_bare_soil_sum', 'total_precipitation_sum', 
               'temperature_2m_max', 'wind_speed_10m']
    ws_data.append(headers)
    
    # 设置列宽
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        if header == 'date':
            ws_data.column_dimensions[col_letter].width = 15  # 日期列宽
        else:
            ws_data.column_dimensions[col_letter].width = 22  # 数据列宽
    
    # 生成15行空数据（带格式和验证）
    today = datetime.now()
    for i in range(1, 16):
        # 日期示例（YYYY-MM-DD格式）
        date_cell = ws_data[f'A{i+1}']
        date_cell.value = f"{(today + timedelta(days=i-1)).strftime('%Y-%m-%d')}"
        date_cell.number_format = 'yyyy-mm-dd'
        
        # 为数值列添加数据验证（必须为数字）
        for col_idx in range(2, 6):
            col_letter = get_column_letter(col_idx)
            cell = ws_data[f'{col_letter}{i+1}']
            
            # 创建数据验证（只能输入数字）
            dv = DataValidation(type="decimal", operator="greaterThan", formula1="-1000")
            dv.error = '请输入有效的数值！'
            dv.errorTitle = '输入错误'
            ws_data.add_data_validation(dv)
            dv.add(cell)
    
    # 添加必填项提示
    ws_data['A18'] = "⚠️ 注意：请填写完整15天的连续数据，不可留空"
    ws_data['A18'].font = openpyxl.styles.Font(color="FF0000", bold=True)
    
    # ===== 工作表2: 填写指南 =====
    ws_guide['A1'] = "数据填写指南"
    ws_guide['A1'].font = openpyxl.styles.Font(size=16, bold=True)
    
    ws_guide['A3'] = "字段说明："
    ws_guide['A3'].font = openpyxl.styles.Font(bold=True)
    
    field_descriptions = {
        'date': '日期 (格式: YYYY-MM-DD，如2025-06-01)',
        'evaporation_from_bare_soil_sum': '裸土蒸发总量 (单位: mm)',
        'total_precipitation_sum': '总降水量 (单位: mm)',
        'temperature_2m_max': '2米高度最高温度 (单位: °C)',
        'wind_speed_10m': '10米高度风速 (单位: m/s)'
    }
    
    row = 4
    for field, desc in field_descriptions.items():
        ws_guide[f'A{row}'] = field
        ws_guide[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        ws_guide[f'B{row}'] = desc
        row += 1
    
    ws_guide['A7'] = "填写要求："
    ws_guide['A7'].font = openpyxl.styles.Font(bold=True)
    ws_guide['B7'] = "1. 必须提供连续15天的完整数据"
    ws_guide['B8'] = "2. 日期需按升序排列"
    ws_guide['B9'] = "3. 数值列不可留空，必须为数字"
    
    # 调整指南工作表列宽
    ws_guide.column_dimensions['A'].width = 30
    ws_guide.column_dimensions['B'].width = 60
    
    # 保存到内存流
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

# Streamlit 主界面
def run_forecast_module():
    st.title("🌧️ 洪水预报模块")
    st.write("上传最新气象数据（Excel 或 CSV），进行未来月径流预测。")
    
    # 生成Excel模板
    excel_buffer = create_excel_template()
    
    # 添加Excel模板下载按钮
    st.download_button(
        label="📊 下载Excel模板",
        data=excel_buffer,
        file_name="data_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="下载Excel格式的数据模板（含填写指南）"
    )
    
    # 添加CSV模板下载按钮
    csv_template = """date,evaporation_from_bare_soil_sum,total_precipitation_sum,temperature_2m_max,wind_speed_10m
YYYY-MM-DD,,,,
YYYY-MM-DD,,,,
YYYY-MM-DD,,,,
YYYY-MM-DD,,,,
YYYY-MM-DD,,,,
YYYY-MM-DD,,,,
YYYY-MM-DD,,,,
YYYY-MM-DD,,,,
YYYY-MM-DD,,,,
YYYY-MM-DD,,,,
YYYY-MM-DD,,,,
YYYY-MM-DD,,,,
YYYY-MM-DD,,,,
YYYY-MM-DD,,,,
YYYY-MM-DD,,,,
"""
    st.download_button(
        label="📄 下载CSV模板",
        data=csv_template,
        file_name="data_template.csv",
        mime="text/csv",
        help="下载CSV格式的数据模板"
    )
    
    # 添加填写提示
    st.info("""
    💡 注意事项：
    1. 请确保填写连续15天的完整数据
    2. 日期格式必须为YYYY-MM-DD（如2025-06-01）
    3. 所有数值列需填写有效数字（如2.5、10.3）
    """)

    # 手动输入 or 文件上传
    manual_input = st.checkbox("手动输入数据")
    df = None

    if manual_input:
        text = st.text_area("输入格式：date,evaporation_from_bare_soil_sum,total_precipitation_sum,temperature_2m_max,wind_speed_10m")
        if text:
            try:
                df = pd.read_csv(StringIO(text)) if ',' in text else pd.read_csv(StringIO(text), sep="\t")
                st.success("✅ 数据读取成功")
                st.dataframe(df.head())
            except Exception as e:
                st.error(f"❌ 数据读取失败：{e}")
                return
        else:
            st.warning("请输入数据")
            return
    else:
        uploaded = st.file_uploader("上传 CSV 或 Excel 文件", type=["csv", "xlsx"])
        if uploaded:
            try:
                df = pd.read_csv(uploaded) if uploaded.name.endswith("csv") else pd.read_excel(uploaded)
                st.success("✅ 文件读取成功")
                st.dataframe(df.head())
            except Exception as e:
                st.error(f"❌ 文件读取失败：{e}")
                return
        else:
            st.warning("请上传数据文件")
            return

    # 检查数据完整性
    if not set(['date'] + DATA_COLUMNS).issubset(df.columns):
        st.error(f"❌ 数据缺失必要列，请确保包含：date + {DATA_COLUMNS}")
        return

    df = df.dropna()
    df['date'] = pd.to_datetime(df['date'])
    features = df[DATA_COLUMNS].values
    dates = df['date'].values

    if len(features) < HISTORY_DAYS:
        st.error(f"❌ 数据长度不足 {HISTORY_DAYS} 天")
        return

    model = load_model()
    last_history = features[-HISTORY_DAYS:]
    last_date = pd.to_datetime(dates[-1])
    predictions, pred_dates = [], []

    for i in range(FORECAST_DAYS):
        input_tensor = torch.tensor(np.expand_dims(last_history, axis=0), dtype=torch.float32)
        with torch.no_grad():
            output = model(input_tensor)
            prediction = output.numpy()[0, -1]
            predictions.append(prediction)
        new_input = last_history[-1]  # 简化处理：用最后一行输入复制
        last_history = np.vstack([last_history[1:], new_input])
        pred_dates.append(last_date + timedelta(days=i+1))

    # 展示结果
    result_df = pd.DataFrame({
        'date': pred_dates,
        'predicted_runoff': predictions
    })
    st.success("✅ 预测完成")
    
    # 显示预测结果图表
    st.subheader("径流预测趋势")
    st.line_chart(result_df.set_index('date'))
    
    st.dataframe(result_df)

    # 下载
    st.download_button("📥 下载预测结果", data=result_df.to_csv(index=False).encode('utf-8'), file_name="direct_forecast.csv")

# 运行页面
if __name__ == "__main__":
    run_forecast_module()
